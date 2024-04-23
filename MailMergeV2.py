from mailmerge import MailMerge
import openpyxl
import os
import pandas as pd

## FOR TEMPLATE A ##

table_dict= {}
row_list = []
row_group = []
col_list = []
header_list = ()
project_list = []
mailing_dict = {}

## Excel Parsing
def excel_parsing(excel_directory):
    bid = openpyxl.load_workbook(excel_directory, data_only=True)
    ws = bid.active
    count = 0
    row_count = 0
    for sheet in bid.worksheets:
        for row in sheet.iter_rows():
            for entry in row:
                try:
                    if '001.' in entry.value:
                        if count%3 == 0:
                            row_count = row_count + 1
                        count = count + 1
                        row_list.append(entry.row)
                        name = str(entry.offset(row=1, column=0).value)
                        address1 = str(entry.offset(row=2, column=0).value)
                        address2 = str(entry.offset(row=3, column=0).value)
                        if entry.offset(row=0, column=1).value == None:
                            lot = ""
                        else:
                            lot = str(entry.offset(row=0, column=1).value)
                        block = str(entry.offset(row=0, column=2).value)
                        value = str(entry.offset(row=3, column=15).value)
                        project = str(entry.offset(row=0, column=17).value)
                        project_stripped = project.replace(" ","")
                        block_stripped = " ".join(block.split())
                        lot_stripped = " ".join(lot.split())
                        table_dict[count] = {}
                        table_dict[count]['Pin'] = entry.value
                        table_dict[count]['Name'] = name
                        table_dict[count]['Address1'] = address1
                        table_dict[count]['Address2'] = address2
                        ##need to clean up lot and blocks, sometimes no lot number or worded strangely, clear whitespaces
                        table_dict[count]['Description'] = 'Lot {} Block {}'.format(lot_stripped, block_stripped)
                        table_dict[count]['Project'] = project_stripped
                        table_dict[count]['Value'] = value
                        table_dict[count]['mailing_row'] = row_count
                        if count == 1:
                            table_dict[count]['mailing_index'] = "col1"
                        else:
                            if table_dict[count - 1]['mailing_index'] == "col1":
                                table_dict[count]['mailing_index'] = "col2"
                            if table_dict[count - 1]['mailing_index'] == "col2":
                                table_dict[count]['mailing_index'] = "col3"
                            if table_dict[count - 1]['mailing_index'] == "col3":
                                table_dict[count]['mailing_index'] = "col1"
                        
                        
                except (AttributeError, TypeError):
                    continue
    start_row = row_list[0]
    table_headers = start_row - 4
    headers = list(ws.iter_rows(max_col=ws.max_column, min_row =table_headers, max_row=table_headers, values_only=True))
    print("Count:{}".format(count))
    return table_dict

def table_sorting():
    for i in table_dict:
        if table_dict[i]['mailing_row'] not in row_group:
            row_group.append(table_dict[i]['mailing_row'])
    print(row_group)

    for k in row_group:
        row_id = k
        mailing_dict[row_id] = {}
        for t in table_dict:
            if table_dict[t]['mailing_row'] == row_id:
                if table_dict[t]['mailing_index'] == "col1":
                    address = table_dict[t]['Address1']
                    mailing_dict[row_id]["column1"] = "{}\n{}\n{}".format(table_dict[t]['Name'], table_dict[t]['Address1'],table_dict[t]['Address2'])
                if table_dict[t]['mailing_index'] == "col2":
                    mailing_dict[row_id]["column2"] = "{}\n{}\n{}".format(table_dict[t]['Name'], table_dict[t]['Address1'],table_dict[t]['Address2'])
                if table_dict[t]['mailing_index'] == "col3":
                    mailing_dict[row_id]["column3"] = "{}\n{}\n{}".format(table_dict[t]['Name'], table_dict[t]['Address1'],table_dict[t]['Address2'])
    print(mailing_dict)
    del i
    del k
    del t

## Mail Merge
def mailmerge(template_directory, folder_locations):
    template = template_directory
    document = MailMerge(template)
    print_count = 0
    document_count = 0
    for key in table_dict:
        if table_dict[key]['Project'] not in project_list:
            project_list.append(table_dict[key]['Project'])
    for i in project_list:
        template = template_directory
        document = MailMerge(template)
        merge_list = []
        for key in table_dict:
            if table_dict[key]['Project'] == i:
                print_count = print_count + 1
                merge_list.append({'Pin': table_dict[key]['Pin'], 'Name': table_dict[key]['Name'], 'Description' : table_dict[key]['Description']})
        document.merge_templates(merge_list, separator='page_break')
        export_path = folder_locations[1]
        docx_file_name = '{}.docx'.format(i)
        print("Exporting {}...".format(docx_file_name))
        document.write(os.path.join(export_path, docx_file_name))
        document_count = document_count + 1
        merge_list.clear()
        document.close()
        del document
    print("{} Documents with a total of {} pages created in Documents".format(document_count, print_count))

def mailing_labels(folder_locations):
    template2 = r"C:\Users\soren.peterson\Desktop\Tempshapes\2024_04_16\table_Test.docx"
    document2 = MailMerge(template2)
    print_count = 0
    document_count = 0
    for i in row_group:
        template = template_directory
        document2 = MailMerge(template2)
        merge_list = []
        merge_list.append({'col1': mailing_dict[i]['column1'], 'col2': mailing_dict[i]['column2'], 'col3' : mailing_dict[i]['column3']})
        document2.merge_rows('col1',
                    merge_list)
        export_path = folder_locations[1]
        docx_file_name = '{}_t.docx'.format(i)
        print("Exporting {}...".format(docx_file_name))
        document2.write(os.path.join(export_path, docx_file_name))
        document_count = document_count + 1
        merge_list.clear()
        document2.close()
        del document2
    print("{} Documents with a total of {} pages created in Documents".format(document_count, print_count))

## Create output folders
def create_project(base_folder):
        output = os.path.join(base_folder, "MailMerge_Export")
        if not os.path.exists(output):
            os.makedirs(output)
            print("Base directory created")
        else:
            print("Base directory already exists")
        output2 = os.path.join(output, "Documents")
        if not os.path.exists(output2):
            os.makedirs(output2)
            print("Document directory created")
        else:
            print("Document directory already exists")
        output3 = os.path.join(output, "Tables")
        if not os.path.exists(output3):
            os.makedirs(output3)
            print("Table directory created")
        else:
            print("Table directory already exists")
        output_folders = [output, output2, output3]
        return output_folders

## Write dictionary to table for mailing labels
def write_excelfile(project_list, folder_locations):
    df = pd.DataFrame.from_dict(table_dict, orient='index')
    excel_count = 0
    for i in project_list:
        query_string = 'Project == "{}"'.format(i)
        filter = df.query(query_string) 
        excel_file_name = "{}.xlsx".format(i)
        filter.to_excel(os.path.join(folder_locations[2], excel_file_name))
        excel_count = excel_count + 1
    print("{} Excel files written in Tables".format(excel_count))

if __name__ == '__main__':
    print("*Excel Mail Merge* \nExcel files must follow one of the assessment templates to be compatible \nNotices must be in .docx format, if other, open the file and save as a .docx")
    excel_directory = input("Enter excel location: ")
    template_directory = input("Enter Docx template location: ")
    export_directory = input("Enter export folder: ")
    excel_parsing(excel_directory)
    table_sorting()
    folder_locations = create_project(export_directory)
    mailmerge(template_directory, folder_locations)
    mailing_labels(folder_locations)
    write_excelfile(project_list, folder_locations)
    project_folder = repr(folder_locations[0])
    print(project_folder)
